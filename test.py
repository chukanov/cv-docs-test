import os
import base64
import requests
import json
import openpyxl
import datetime

# === Настройки ===
TEST_DATA_DIR = "test_data2"
EXCEL_FILE = os.path.join(TEST_DATA_DIR, "test.xlsx")
API_URL = "http://localhost:8000/v1/verifyStudentsData"
RESULTS_FILE = "test_results.xlsx"
AUTH_TOKEN = "05cf7d9e-b8fe"

# === Функция кодирования файла в Base64 ===
def encode_file_to_base64(file_path):
    with open(file_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")

# === Определение MIME-типа по расширению ===
MIME_TYPES = {
    "jpeg": "image/jpeg",
    "jpg": "image/jpeg",
    "png": "image/png",
    "pdf": "application/pdf"
}

def get_mime_type(filename):
    ext = filename.split(".")[-1].lower()
    return MIME_TYPES.get(ext, "application/octet-stream")

# === Чтение Excel (XLSX) ===
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active  # Берем первый лист

test_results = []

# Читаем заголовки (проверка регистрозависимости)
ex_headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value is not None]
print("Заголовки колонок в файле:", ex_headers)

for row in ws.iter_rows(min_row=2, values_only=True):
    # Пропускаем пустые строки
    if not any(row):
        continue  

    if len(row) < len(ex_headers):
        print(f"⚠️ Пропущены данные в строке {row}. Пропускаем.")
        continue

    row_data = {ex_headers[i]: row[i] for i in range(len(ex_headers))}

    last_name = row_data["lastname"]
    first_name = row_data["firstname"]
    sur_name = row_data["surname"]
    gender = row_data["gender"]

    birth_date = row_data["birthdate"]
    if isinstance(birth_date, (datetime.datetime, datetime.date)):  
        birth_date = birth_date.strftime("%Y-%m-%d")

    diploma_country = row_data["diplomacountry"]
    education_level = row_data["educationlevel"]
    diploma_last_name = row_data.get("diplomalastname", "")

    user_folder = os.path.join(TEST_DATA_DIR, f"{last_name} {first_name} {sur_name}")

    if not os.path.isdir(user_folder):
        print(f"⚠️ Папка {user_folder} не найдена. Пропускаем.")
        continue

    # Ищем файлы паспорта, диплома и свидетельства о браке
    user_files = {}
    for file_type, file_prefix in {"passport": "p", "diploma": "d", "marriageCertificate": "s"}.items():
        for ext in MIME_TYPES.keys():
            file_path = os.path.join(user_folder, f"{file_prefix}.{ext}")
            if os.path.exists(file_path):
                user_files[file_type] = {
                    "fileName": os.path.basename(file_path),
                    "mimeType": get_mime_type(file_path),
                    "data": encode_file_to_base64(file_path)
                }
                break
        else:
            if file_type != "marriageCertificate":  # Свидетельство о браке не обязательно
                print(f"⚠️ Файл {file_prefix} не найден в {user_folder}. Пропускаем.")
                continue

    # Формируем JSON-запрос
    payload = {
        "userInfo": {
            "firstName": first_name,
            "surName": sur_name,
            "lastName": last_name,
            "gender": gender,
            "birthDate": birth_date,
            "diplomaCountry": diploma_country,
            "educationLevel": education_level,
            "diplomaLastName": diploma_last_name
        },
        "userFiles": user_files
    }

    # Заголовки запроса с авторизацией
    headers = {
        "Authorization": f"Bearer {AUTH_TOKEN}",
        "Content-Type": "application/json"
    }

    # Отправка запроса
    try:
        response = requests.post(API_URL, json=payload, timeout=30, headers=headers)
        response_json = response.json()
        status = response_json.get("status", "error")
        if status == "success":
            result = "✅ Passed"
        else:
            raw_resp = json.dumps(response_json, ensure_ascii=False)
            result = f"❌ Failed ({raw_resp})"
    except Exception as e:
        result = f"❌ Failed response: {str(e)}"

    # Логирование результата
    print(f"🔹 {last_name} {first_name} {sur_name}: {result}")

    # Сохранение результата
    test_results.append([last_name, first_name, sur_name, result])

# === Запись результатов в Excel (XLSX) ===
wb_results = openpyxl.Workbook()
ws_results = wb_results.active
ws_results.append(["Фамилия", "Имя", "Отчество", "Результат"])  # Заголовки

for row in test_results:
    ws_results.append(row)

wb_results.save(RESULTS_FILE)
print("\n📜 Тестирование завершено! Результаты сохранены в", RESULTS_FILE)
